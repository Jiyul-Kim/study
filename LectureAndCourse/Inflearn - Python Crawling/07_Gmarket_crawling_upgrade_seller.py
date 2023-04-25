import requests
from bs4 import BeautifulSoup
import openpyxl

excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.column_dimensions['B'].width = 100
excel_sheet.column_dimensions['C'].width = 30
excel_sheet.column_dimensions['D'].width = 30
excel_sheet.column_dimensions['E'].width = 100
excel_sheet.append(['랭킹', '제목', '가격', '판매자', '링크'])


url_list = []

url2 = 'https://corners.gmarket.co.kr/Bestsellers'
res2 = requests.get(url2)
soup2 = BeautifulSoup(res2.content, 'html.parser')

sheet_names = soup2.find_all('ul', {'class':'by-group'})
for sheet_name in sheet_names:
    sheet_name = sheet_name.get_text().replace('/', '')
    print(sheet_name)
    b = sheet_name.split()
    print(b)

for index in range(1, 13):
    if index == 1:
        url = "https://corners.gmarket.co.kr/Bestsellers"    
    if index <= 9:
        index = '0' + str(index)
    url = "https://corners.gmarket.co.kr/Bestsellers?viewType=G&groupCode=G" + str(index)

    url_list.append(url)

    for index3, i in enumerate(b):
        if index3 == 0:
            excel_file.create_sheet(i)
            del excel_file['Sheet']
            continue
        excel_file.create_sheet(i)     

        url = 'https://corners.gmarket.co.kr/Bestsellers'
        res = requests.get(url)
        soup = BeautifulSoup(res.content, 'html.parser')

            
        bestlists = soup.select_one('div.best-list')
        prodects = bestlists.select('ul > li')

        for index2, prodect in enumerate(prodects):
            if index2 > 10:
                title = prodect.select_one('a.itemname').get_text()
                price = prodect.select_one('div.item_price').get_text()
                # print(str(index2 + 1)+ '.',  title.get_text(), price.get_text(),end="")
                
                link = prodect.select_one('a')['href']
                res2 = requests.get(link)
                soup2 = BeautifulSoup(res2.content, 'html.parser')
                
                seller = soup2.select_one('span.text__seller > a').get_text()


                excel_sheet.append([str(index2 + 1) + '.', title, price , seller, link])
                # excel_sheet.cell(row = index + 2, column = 5).hyperlink = link
                
                print(str(index2 + 1) + '.'
                    , title
                    , price
                    , seller
                    , link)
                break
        excel_file.save('Gmarket_Crawling3.xlsx')
        excel_file.close()